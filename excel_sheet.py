import openpyxl as xl

from utils import excel_file_path


class WorkBook:
    """Basic workbook class, provides methods to work with sheet's cell values"""

    def __init__(self, file_path, sheet_name=None):
        self.__file_path = file_path
        self.__wb = xl.load_workbook(filename=self.__file_path)
        self.__sheet_name = sheet_name if sheet_name else self.__wb.sheetnames[0]
        self.__sheet = self.__wb[self.__sheet_name]

    def __get_row_values(self, row_id):
        column_id = 1
        row_values = []

        while True:
            cell_value = self.__sheet.cell(row=row_id, column=column_id).value
            row_values.append(cell_value)
            column_id += 1
            if cell_value is None:
                break

        return row_values[:-1]

    def get_sheet_values(self):
        sheet_data = []
        for row_id in range(2, self.__sheet.max_row + 1):
            sheet_data.append(self.__get_row_values(row_id=row_id))

        return sheet_data


wb = WorkBook(file_path=excel_file_path)
wb_data = wb.get_sheet_values()

if __name__ == '__main__':
    print(wb_data)

